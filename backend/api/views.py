import json
import base64
import os
from datetime import timedelta
from django.utils import timezone
from django.http import JsonResponse, FileResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from api.models import User
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from django.db.models import Q
import tempfile

@csrf_exempt
def login_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            employee_code = data.get('employee_code').upper()
            password = data.get('password')

            if not employee_code or not password:
                return JsonResponse({'success': False, 'message': 'Employee code and password are required.'}, status=400)
            
            user = User.objects.filter(employee_code=employee_code).first()

            if user and user.verify_password(password):
                if not user.is_active:
                    return JsonResponse({'success': False, 'message': 'Account is disabled.'}, status=403)
                
                return JsonResponse({
                    'success': True, 
                    'message': 'Login successful.',
                    'user': {
                        'employee_code': user.employee_code,
                        'full_name': user.full_name,
                        'email': user.email,
                        'role': user.role,
                        'is_initial_password': user.is_initial_password if user.role != 'admin' else False
                    }
                })
            else:
                return JsonResponse({'success': False, 'message': 'Invalid employee code or password.'}, status=401)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@csrf_exempt
def add_user(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            employee_code = data.get('employee_code').upper()
            email_id = data.get('email_id').lower()
            full_name = data.get('full_name')
            password = data.get('password')
            role = data.get('role', 'Viewer').lower() # defaults to viewer/user mapping

            if not all([employee_code, email_id, full_name, password]):
                return JsonResponse({'success': False, 'message': 'All fields are required.'}, status=400)

            # Check if user already exists
            if User.objects.filter(employee_code=employee_code).exists():
                return JsonResponse({'success': False, 'message': 'Employee code already exists.'}, status=400)
            if User.objects.filter(email=email_id).exists():
                return JsonResponse({'success': False, 'message': 'Email address already exists.'}, status=400)

            # Map the frontend "User" role to backend "user" or "viewer"
            # Based on models.py: Admin, User, Viewer
            backend_role = 'user'
            if role == 'admin':
                backend_role = 'admin'
            elif role == 'viewer':
                backend_role = 'viewer'

            user = User(
                employee_code=employee_code,
                email=email_id,
                full_name=full_name,
                role=backend_role,
                is_active=True
            )
            user.set_password(password)
            user.save()

            return JsonResponse({
                'success': True,
                'message': 'User created successfully.',
                'user': {
                    'employee_code': user.employee_code,
                    'full_name': user.full_name,
                    'email': user.email,
                    'role': user.role
                }
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@csrf_exempt
def get_users(request):
    if request.method == 'GET':
        try:
            users_query = User.objects.filter(is_deleted=False).order_by('created_at')
            users_list = []
            for u in users_query:
                users_list.append({
                    'employee_code': u.employee_code,
                    'full_name': u.full_name,
                    'email': u.email,
                    'role': u.role,
                    'is_active': u.is_active,
                    'created_at': u.created_at.isoformat()
                })
            return JsonResponse({'success': True, 'users': users_list})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
            
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@csrf_exempt
def delete_user(request, employee_code):
    if request.method == 'DELETE':
        try:
            user = User.objects.get(employee_code=employee_code, is_deleted=False)
            user.is_deleted = True
            user.save()
            return JsonResponse({'success': True, 'message': 'User deleted successfully.'})
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'User not found.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@csrf_exempt
def update_user(request, employee_code):
    if request.method == 'PATCH':
        try:
            data = json.loads(request.body)
            user = User.objects.get(employee_code=employee_code, is_deleted=False)

            if 'full_name' in data:
                user.full_name = data['full_name']
            if 'email_id' in data:
                data['email_id'] = data['email_id'].lower()
                # Check email uniqueness (exclude self)
                if User.objects.filter(email=data['email_id']).exclude(employee_code=employee_code).exists():
                    return JsonResponse({'success': False, 'message': 'Email address already in use.'}, status=400)
                user.email = data['email_id']
            if 'role' in data:
                user.role = data['role'].lower()
            if 'is_active' in data:
                user.is_active = data['is_active']
            if 'password' in data and data['password']:
                user.set_password(data['password'])
                user.is_initial_password = True

            user.save()
            return JsonResponse({
                'success': True,
                'message': 'User updated successfully.',
                'user': {
                    'employee_code': user.employee_code,
                    'full_name': user.full_name,
                    'email': user.email,
                    'role': user.role,
                    'is_active': user.is_active,
                }
            })
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'User not found.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@csrf_exempt
def download_file(request):
    """
    Downloads a file given a base64-encoded file_path.
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
        
    encoded_path = request.GET.get('file_path')
    if not encoded_path:
        raise Http404("No file path provided")
    
    try:
        file_path = base64.b64decode(encoded_path).decode()
        if os.path.exists(file_path):
            response = FileResponse(open(file_path, 'rb'), as_attachment=True)
            return response
        else:
            raise Http404(f"File not found: {file_path}")
    except Exception as e:
        raise Http404(f"Invalid file path: {str(e)}")

@csrf_exempt
def update_initial_setup(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            employee_code = data.get('employee_code').upper() if data.get('employee_code') else None
            new_password = data.get('new_password')
            q1 = data.get('q1')
            a1 = data.get('a1')

            if not all([employee_code, new_password, q1, a1]):
                return JsonResponse({'success': False, 'message': 'All fields are required.'}, status=400)

            user = User.objects.filter(employee_code=employee_code).first()
            if not user:
                return JsonResponse({'success': False, 'message': 'User not found.'}, status=404)

            # Update password and security questions
            user.set_password(new_password)
            user.security_q1 = q1
            user.security_a1 = a1.lower().strip()
            user.is_initial_password = False
            user.save()

            return JsonResponse({'success': True, 'message': 'Password and security questions updated successfully.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@csrf_exempt
def forgot_password_request(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email', '').lower()
            if not email:
                return JsonResponse({'success': False, 'message': 'Email is required.'}, status=400)
            
            user = User.objects.filter(email=email).first()
            if not user:
                return JsonResponse({'success': False, 'message': 'User with this email not found.'}, status=404)
            
            if not user.security_q1:
                return JsonResponse({'success': False, 'message': 'Security question not set for this account. Please contact admin.'}, status=400)
            
            return JsonResponse({'success': True, 'question': user.security_q1})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@csrf_exempt
def forgot_password_reset(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email', '').lower()
            answer = data.get('answer', '')
            new_password = data.get('new_password')

            if not all([email, answer, new_password]):
                return JsonResponse({'success': False, 'message': 'All fields are required.'}, status=400)
            
            user = User.objects.filter(email=email).first()
            if not user:
                return JsonResponse({'success': False, 'message': 'User not found.'}, status=404)
            
            if user.security_a1.lower() != answer.lower().strip():
                return JsonResponse({'success': False, 'message': 'Incorrect security answer.'}, status=401)
            
            user.set_password(new_password)
            user.is_initial_password = False
            user.save()

            return JsonResponse({'success': True, 'message': 'Password reset successfully.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@csrf_exempt
def get_stats(request):
    if request.method == 'GET':
        try:
            from .models import FileProcessingLog
            total_users = User.objects.filter(is_deleted=False).count()
            active_users = User.objects.filter(is_deleted=False, is_active=True).count()
            inactive_users = User.objects.filter(is_deleted=False, is_active=False).count()
            
            total_processed = FileProcessingLog.objects.count()
            
            now = timezone.now()
            today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            processed_today = FileProcessingLog.objects.filter(processed_at__gte=today_start).count()

            # Aggregate granular metrics
            total_sw_contra = 0
            total_final_contra = 0
            total_sw_return = 0
            total_final_return = 0
            
            import re
            def parse_metric(text, keys):
                if not text: return 0
                total = 0
                for key in keys:
                    m = re.search(fr'{key}:\s*(\d+)', str(text).lower())
                    if m:
                        total += int(m.group(1))
                return total

            for log in FileProcessingLog.objects.all():
                total_sw_contra += parse_metric(log.software_count, ['inb_trf', 'sis_con'])
                total_final_contra += parse_metric(log.final_count, ['inb_trf', 'sis_con'])
                total_sw_return += parse_metric(log.software_count, ['return'])
                total_final_return += parse_metric(log.final_count, ['return'])
            
            # Efficiency Percentages
            contra_efficiency = (total_sw_contra / total_final_contra * 100) if total_final_contra > 0 else 0
            return_efficiency = (total_sw_return / total_final_return * 100) if total_final_return > 0 else 0
            
            admin_count = User.objects.filter(is_deleted=False, role='admin').count()
            staff_count = User.objects.filter(is_deleted=False, role='staff').count()
            regular_user_count = User.objects.filter(is_deleted=False, role='user').count()
            
            now = timezone.now()
            today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            week_start = today_start - timedelta(days=7)
            
            daily_active = User.objects.filter(is_deleted=False, last_login__gte=today_start).count()
            weekly_active = User.objects.filter(is_deleted=False, last_login__gte=week_start).count()
            new_users_week = User.objects.filter(is_deleted=False, created_at__gte=week_start).count()
            
            # Recent logins list
            recent_logins = []
            for u in User.objects.filter(is_deleted=False, last_login__isnull=False).order_by('-last_login')[:5]:
                recent_logins.append({
                    'full_name': u.full_name,
                    'employee_code': u.employee_code,
                    'last_login': u.last_login.isoformat()
                })

            # Monthly Trends (Last 6 Months)
            monthly_trends = []
            for i in range(5, -1, -1):
                # Calculate the first day of that month
                # Use a more robust month subtraction
                year = now.year
                month = now.month - i
                while month <= 0:
                    month += 12
                    year -= 1
                
                m_logs = FileProcessingLog.objects.filter(
                    processed_at__year=year,
                    processed_at__month=month
                )
                
                m_sw_contra = 0
                m_final_contra = 0
                m_sw_return = 0
                m_final_return = 0
                
                from .helpers import parse_metric
                for l in m_logs:
                    m_sw_contra += parse_metric(l.software_count, ['inb_trf', 'sis_con'])
                    m_final_contra += parse_metric(l.final_count, ['inb_trf', 'sis_con'])
                    m_sw_return += parse_metric(l.software_count, ['return'])
                    m_final_return += parse_metric(l.final_count, ['return'])
                
                month_name = timezone.datetime(year, month, 1).strftime("%b %y")
                monthly_trends.append({
                    "name": month_name,
                    "contra": round((m_sw_contra / m_final_contra * 100), 2) if m_final_contra > 0 else 0,
                    "return": round((m_sw_return / m_final_return * 100), 2) if m_final_return > 0 else 0,
                    "files": m_logs.count()
                })

            return JsonResponse({
                'success': True,
                'stats': {
                    'total_users': total_users,
                    'active_users': active_users,
                    'inactive_users': inactive_users,
                    'total_processed': total_processed,
                    'processed_today': processed_today,
                    'total_sw_contra': total_sw_contra,
                    'total_final_contra': total_final_contra,
                    'total_sw_return': total_sw_return,
                    'total_final_return': total_final_return,
                    'contra_efficiency': round(contra_efficiency, 2),
                    'return_efficiency': round(return_efficiency, 2),
                    'roles': {
                        'admin': admin_count,
                        'staff': staff_count,
                        'user': regular_user_count
                    },
                    'activity': {
                        'daily_active': daily_active,
                        'weekly_active': weekly_active,
                        'new_users_week': new_users_week
                    },
                    'recent_logins': recent_logins,
                    'monthly_trends': monthly_trends
                }
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
def get_processing_logs(request):
    """Retrieve file processing logs for the statistics dashboard."""
    if request.method == 'GET':
        from .models import FileProcessingLog
        # Retrieve the latest 100 logs
        logs = FileProcessingLog.objects.all().order_by('-processed_at')[:100]
        data = []
        from .helpers import parse_metric
        for l in logs:
            total = l.total_entries
            # sw_matched = non-empty type count
            sw_matched = parse_metric(l.software_count)
            # contra = inb_trf + sis_con
            contra_count = parse_metric(l.software_count, ['inb_trf', 'sis_con'])
            # return = return
            return_count = parse_metric(l.software_count, ['return'])
            
            sw_matched_pct = round((sw_matched / total * 100), 2) if total > 0 else 0
            # team_contribution = empty type count (total - matched)
            team_contrib_pct = round(((total - sw_matched) / total * 100), 2) if total > 0 else 0
            
            # Final categorized total for contra/return percentages
            final_contra_count = parse_metric(l.final_count, ['inb_trf', 'sis_con'])
            final_return_count = parse_metric(l.final_count, ['return'])
            
            contra_pct = round((contra_count / final_contra_count * 100), 2) if final_contra_count > 0 else 0
            return_pct = round((return_count / final_return_count * 100), 2) if final_return_count > 0 else 0

            data.append({
                "id": l.id,
                "user_name": l.user_name,
                "file_name": l.file_name,
                "bank_name": l.bank_name or "Unknown",
                "processed_at": l.processed_at.strftime("%Y-%m-%d %H:%M:%S"),
                "total_entries": total,
                "sw_matched_pct": round(sw_matched_pct, 2),
                "team_contrib_pct": round(team_contrib_pct, 2),
                "contra_pct": round(contra_pct, 2),
                "return_pct": round(return_pct, 2),
            })
        return JsonResponse(data, safe=False)
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
def save_incentive(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            date = data.get('date')
            client_name = data.get('client_name')
            for_user = data.get('for_user')
            client_type = data.get('client_type')
            shares = data.get('shares')
            
            # Recalculate total_amount from shares for integrity (handles Indian comma format)
            calculated_total = 0
            if isinstance(shares, list):
                for s in shares:
                    amt_str = str(s.get('amount', '0')).replace(',', '')
                    try:
                        calculated_total += float(amt_str)
                    except ValueError:
                        pass
            
            total_amount = calculated_total
            total_reward = data.get('total_reward')
            employee_code = data.get('employee_code')

            if not all([date, client_name, client_type, shares, total_reward]):
                return JsonResponse({'success': False, 'message': 'Missing required fields.'}, status=400)

            user = User.objects.filter(employee_code=employee_code).first() if employee_code else None

            from .models import Incentive
            incentive = Incentive(
                date=date,
                for_user=for_user,
                client_name=client_name,
                client_type=client_type,
                shares=shares,
                total_amount=total_amount,
                total_reward=total_reward,
                created_by=user
            )
            incentive.save()

            return JsonResponse({'success': True, 'message': 'Incentive record saved successfully.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
def get_incentives(request):
    if request.method == 'GET':
        try:
            from .models import Incentive
            records = Incentive.objects.all().order_by('-created_at')
            data = []
            for r in records:
                data.append({
                    'id': r.id,
                    'date': str(r.date),
                    'for_user': r.for_user,
                    'client_name': r.client_name,
                    'client_type': r.client_type,
                    'shares': r.shares,
                    'total_amount': float(r.total_amount),
                    'total_reward': float(r.total_reward),
                    'created_at': r.created_at.strftime("%Y-%m-%d %H:%M:%S")
                })
            return JsonResponse({'success': True, 'records': data})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
def export_incentives(request):
    """
    Exports incentive records to a professionally formatted Excel file.
    Follows a dual-table layout per recipient with specific styling.
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    for_user = request.GET.get('for_user')
    month = request.GET.get('month')
    
    from .models import Incentive
    query = Q()
    if for_user:
        query &= Q(for_user=for_user)
    if month:
        try:
            query &= Q(date__month=int(month))
        except ValueError:
            pass
    
    records = Incentive.objects.filter(query).order_by('date', 'for_user')
    
    if not records.exists():
        return JsonResponse({'success': False, 'message': 'No records found for the given filters.'}, status=404)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    # Define styles
    header_fill_1 = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid") # Cyan (using 00FFFF instead of CCFFFF for better contrast)
    header_fill_2 = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid") # Gold
    border_side = Side(style='thin')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    header_font = Font(bold=True, size=10)
    total_font = Font(bold=True, color="FF0000") # Red totals as in screenshot
    
    # Group by recipient for sheets
    recipients = records.values_list('for_user', flat=True).distinct()
    
    for recipient in recipients:
        sheet_name = str(recipient) if recipient else "General"
        # Sheet names have 31 char limit and no special chars
        sheet_name = "".join([c for c in sheet_name if c.isalnum() or c in " _-"])[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        recipient_records = records.filter(for_user=recipient)
        
        # Get all unique companies for this recipient's records
        all_companies = []
        for r in recipient_records:
            if isinstance(r.shares, list):
                for s in r.shares:
                    comp = s.get('company')
                    if comp and comp not in all_companies:
                        all_companies.append(comp)
        all_companies.sort()
        
        # --- TABLE 1: SHARE AMOUNTS ---
        # Recipient Name header (Red text at the top)
        ws.cell(row=1, column=1, value=str(recipient).upper()).font = Font(bold=True, color="FF0000")
        
        # Headers Table 1
        headers1 = ["SI NO", "DATE", "NAME", "LOAN AMOUNT"] + all_companies + ["CLIENT"]
        ws.append(headers1)
        header_row = 2
        for col, val in enumerate(headers1, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = header_fill_1
            cell.border = border
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            
        # Data Table 1
        curr_row = 3
        total_loan = 0
        for idx, r in enumerate(recipient_records, 1):
            row_data = [
                idx,
                r.date.strftime("%d-%m-%Y"),
                r.client_name,
                float(r.total_amount)
            ]
            total_loan += float(r.total_amount)
            
            # Company-wise principal amounts
            comp_amounts = {s.get('company'): s.get('amount') for s in r.shares} if isinstance(r.shares, list) else {}
            for comp in all_companies:
                val = comp_amounts.get(comp, "")
                if val:
                    try:
                        val = float(str(val).replace(',', ''))
                    except ValueError:
                        val = 0
                row_data.append(val if val else "-")
                
            row_data.append(r.client_type.upper())
            
            # Write data row
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=curr_row, column=col)
                cell.value = val
                cell.border = border
                if col == 4 or (col > 4 and col <= 4+len(all_companies)): # Amount columns
                    cell.number_format = '#,##,##0.00'
                if col == 4:
                    cell.font = Font(bold=True)
            
            curr_row += 1
            
        # Totals Table 1
        total_row_idx = curr_row
        ws.cell(row=total_row_idx, column=3, value="Total").font = total_font
        ws.cell(row=total_row_idx, column=4, value=total_loan).font = total_font
        ws.cell(row=total_row_idx, column=4).number_format = '#,##,##0.00'
        
        # Apply borders to the full total row
        for col in range(1, len(headers1) + 1):
            ws.cell(row=total_row_idx, column=col).border = border
            
        # --- SPACER ---
        curr_row += 4
        
        # --- TABLE 2: INCENTIVE AMOUNTS ---
        headers2 = ["S.NO", "DATE", "NAME", "LOAN AMOUNT"] + all_companies + ["AMOUNT", "SHARE"]
        header_row2 = curr_row
        for col, val in enumerate(headers2, 1):
            cell = ws.cell(row=header_row2, column=col)
            cell.value = val
            cell.fill = header_fill_2
            cell.border = border
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            
        curr_row += 1
        start_data_row2 = curr_row
        total_incentive_overall = 0
        comp_incentive_totals = {comp: 0 for comp in all_companies}
        
        for idx, r in enumerate(recipient_records, 1):
            row_data = [
                idx,
                r.date.strftime("%d-%m-%Y"),
                r.client_name,
                float(r.total_amount)
            ]
            
            # Company-wise incentive amounts
            comp_incentives = {s.get('company'): s.get('incentive') for s in r.shares} if isinstance(r.shares, list) else {}
            for comp in all_companies:
                val = comp_incentives.get(comp, 0)
                try:
                    val_float = float(val) if val else 0
                except (ValueError, TypeError):
                    val_float = 0
                comp_incentive_totals[comp] += val_float
                row_data.append(val_float if val_float else "-")
                
            row_data.append(float(r.total_reward)) # Total Incentive
            row_data.append("") # Share (blank in screenshot)
            total_incentive_overall += float(r.total_reward)
            
            # Write data row
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=curr_row, column=col)
                cell.value = val
                cell.border = border
                if col >= 4:
                    cell.number_format = '#,##,##0.00'
                if col == 4:
                     cell.font = Font(bold=True)
            curr_row += 1
            
        # Totals Table 2
        total_row_idx2 = curr_row
        ws.cell(row=total_row_idx2, column=3, value="Total").font = total_font
        ws.cell(row=total_row_idx2, column=4, value=total_loan).font = total_font
        ws.cell(row=total_row_idx2, column=4).number_format = '#,##,##0.00'
        
        for idx, comp in enumerate(all_companies):
            ws.cell(row=total_row_idx2, column=5+idx, value=comp_incentive_totals[comp]).font = total_font
            ws.cell(row=total_row_idx2, column=5+idx).number_format = '#,##,##0.00'
            
        ws.cell(row=total_row_idx2, column=5+len(all_companies), value=total_incentive_overall).font = total_font
        ws.cell(row=total_row_idx2, column=5+len(all_companies)).number_format = '#,##,##0.00'
        
        # Final set borders for total row
        for col in range(1, len(headers2) + 1):
            ws.cell(row=total_row_idx2, column=col).border = border

        # Adjust column widths
        for col_cells in ws.columns:
            # Avoid error on empty cells
            values = [str(cell.value or "") for cell in col_cells]
            if not values:
                continue
            length = max(len(v) for v in values)
            ws.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 40)

    # Save to response using a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    response = FileResponse(open(tmp_path, 'rb'), as_attachment=True, 
                            filename=f'Incentive_Report_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx')
    return response
