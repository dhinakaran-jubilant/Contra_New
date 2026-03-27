from . import helpers, config
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell
from copy import copy
import pandas as pd

# ── Constants ─────────────────────────────────────────────────────────────────
_ANALYSIS_COL_LETTERS = ["C", "D", "E", "F", "G", "H", "I"]
_ANALYSIS_COL_WIDTH   = 15

_HEADER_FILL   = PatternFill(fill_type="solid", start_color="FF002060", end_color="FF002060")
_HEADER_FONT   = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
_BODY_FONT     = Font(name="Arial", size=10)
_CATEGORY_FONT = Font(name="Calibri", size=10, bold=True)
_THIN_SIDE     = Side(border_style="thin", color="000000")
_THIN_BORDER   = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_CENTER_ALIGN  = Alignment(horizontal="center", vertical="center")
_VCENTER_ALIGN = Alignment(vertical="center")
_WRAP_CENTER   = Alignment(wrap_text=True, horizontal="center", vertical="center")
_WRAP_VCENTER  = Alignment(wrap_text=True, vertical="center")

_NUMBER_FORMAT = "#,##,##0.00"
_DATE_FORMAT   = "DD-MMM-YY"

_DESIRED_WIDTHS = {
    "Sl. No.":     8,
    "Date":        10,
    "MONTH":       10,
    "TYPE":        18,
    "Cheque_No":   12,
    "Category":    35,
    "Description": 50,
    "DR":          15,
    "CR":          15,
    "Balance":     18,
}

_TABLE_STYLES = [
    {
        "selector": "th",
        "props": "background-color: #002060; color: white; border: 1px solid black;",
    },
    {"selector": "td", "props": "border: 1px solid black;"},
]

# ── Private helpers ────────────────────────────────────────────────────────────

def _copy_cell_style(src_cell, dest_cell) -> None:
    """Copy all style attributes from *src_cell* to *dest_cell*."""
    if src_cell.has_style:
        dest_cell.font          = copy(src_cell.font)
        dest_cell.fill          = copy(src_cell.fill)
        dest_cell.border        = copy(src_cell.border)
        dest_cell.number_format = src_cell.number_format
        dest_cell.alignment     = copy(src_cell.alignment)
        dest_cell.protection    = copy(src_cell.protection)


def _apply_merge_style(dest_ws, coord: str) -> None:
    """
    Propagate the top-left cell's border and fill to every cell
    inside a merged region identified by *coord* (e.g. 'A1:C3').
    """
    top_left_addr = coord.split(":")[0]
    top_left      = dest_ws[top_left_addr]
    border        = copy(top_left.border)
    fill          = copy(top_left.fill)
    for row in dest_ws[coord]:
        for cell in row:
            cell.border = border
            cell.fill   = fill


def _set_analysis_col_widths(ws) -> None:
    """Set fixed width for the standard ANALYSIS columns."""
    for letter in _ANALYSIS_COL_LETTERS:
        ws.column_dimensions[letter].width = _ANALYSIS_COL_WIDTH


def _reorder_sheet_first(wb, ws) -> None:
    """Move *ws* to be the first sheet in *wb*."""
    sheets = list(wb._sheets)
    if ws in sheets:
        sheets.insert(0, sheets.pop(sheets.index(ws)))
        wb._sheets = sheets


def _apply_number_format_col(ws, col_idx: int, min_row: int = 1,
                              fmt: str = _NUMBER_FORMAT,
                              check_negative: bool = False) -> None:
    """
    Apply *fmt* to every cell in column *col_idx* starting at *min_row*.
    When *check_negative* is True, also colour negative values red.
    """
    for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row,
                            min_col=col_idx, max_col=col_idx):
        cell = row[0]
        cell.number_format = fmt
        if check_negative and cell.value is not None:
            try:
                val = float(str(cell.value).replace(",", ""))
                if val < 0:
                    cell.font = cell.font.copy(color="FFFF0000")
            except (ValueError, TypeError):
                pass


def _safe_name_from_key(key: str) -> str:
    """Strip 'XNS' prefix from a sheet key and join the rest with '-'."""
    parts = key.split("-")
    if "XNS" in parts:
        parts.remove("XNS")
    return "-".join(parts)


# ── Public API ─────────────────────────────────────────────────────────────────

def copy_sheet_with_style(src_ws, wb_out, new_title: str = "Analysis"):
    """Copy *src_ws* into *wb_out* as a new sheet, preserving all styles."""
    dest_ws = wb_out.create_sheet(title=new_title)

    for row in src_ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            dcell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            _copy_cell_style(cell, dcell)

    for r_idx, r_dim in src_ws.row_dimensions.items():
        if r_dim.height is not None:
            dest_ws.row_dimensions[r_idx].height = r_dim.height

    for col_letter, c_dim in src_ws.column_dimensions.items():
        if c_dim.width is not None:
            dest_ws.column_dimensions[col_letter].width = c_dim.width

    for merged in src_ws.merged_cells.ranges:
        coord = str(merged)
        dest_ws.merge_cells(coord)
        _apply_merge_style(dest_ws, coord)

    _set_analysis_col_widths(dest_ws)
    return dest_ws


def append_sheet_with_style(src_ws, dest_ws, gap_rows: int = 2, col_offset: int = 0):
    """Append *src_ws* data (with styles) after the existing content of *dest_ws*."""
    row_offset = dest_ws.max_row + gap_rows

    for row in src_ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            target_row = row_offset + cell.row - 1
            target_col = cell.column + col_offset
            dcell = dest_ws.cell(row=target_row, column=target_col, value=cell.value)
            _copy_cell_style(cell, dcell)

    for r_idx, r_dim in src_ws.row_dimensions.items():
        if r_dim.height is not None:
            dest_ws.row_dimensions[row_offset + r_idx - 1].height = r_dim.height

    for col_letter, c_dim in src_ws.column_dimensions.items():
        if c_dim.width is None:
            continue
        src_col_idx  = column_index_from_string(col_letter)
        dest_col_idx = src_col_idx + col_offset
        dest_letter  = get_column_letter(dest_col_idx)
        dest_dim     = dest_ws.column_dimensions[dest_letter]
        dest_dim.width = max(dest_dim.width or 0, c_dim.width)

    for merged in src_ws.merged_cells.ranges:
        min_r = merged.min_row + row_offset - 1
        max_r = merged.max_row + row_offset - 1
        min_c = merged.min_col + col_offset
        max_c = merged.max_col + col_offset
        coord = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"
        dest_ws.merge_cells(coord)
        _apply_merge_style(dest_ws, coord)

    return dest_ws


def save_matched_with_styles(
        all_files_to_process: dict,
        acc_name_storage: dict,
        analysis_storage: dict,
        statement_storage: dict,
        highlight_red_positions=None,
        highlight_green_positions=None,
) -> list:
    """
    Write each processed DataFrame to a styled .xlsx file.
    Returns a list of dicts with file_path, file_name, account_name, sheet_name.
    """
    saved_files_info = []
    needs_highlight  = highlight_red_positions is not None or highlight_green_positions is not None

    if needs_highlight:
        light_red_fill   = PatternFill(fill_type="solid", start_color="FFFFCCCC", end_color="FFFFCCCC")
        light_green_fill = PatternFill(fill_type="solid", start_color="FFC6EFCE", end_color="FFC6EFCE")

    print(f"💾 Saving {len(all_files_to_process)} files with styles...")

    for key, df in all_files_to_process.items():
        print(f"   Processing: {key}")

        safe_name    = _safe_name_from_key(key)
        acc_name     = acc_name_storage[key]
        safe_acc     = helpers.sanitize_filename(acc_name)

        download_path = helpers.get_downloads()
        base_dir      = download_path / config.PROCESSED_DIR
        base_dir.mkdir(parents=True, exist_ok=True)
        filename = base_dir / f"{safe_acc}-{safe_name}.xlsx"

        df_out = df.copy()
        df_out["DR"] = pd.to_numeric(df_out["DR"], errors="coerce").abs()
        df_out["CR"] = pd.to_numeric(df_out["CR"], errors="coerce")

        styler = (
            df_out.style
            .set_table_styles(_TABLE_STYLES)
            .set_properties(**{"border": "1px solid black"})
            .set_properties(subset=["Category"], **{"font-weight": "bold"})
            .format({"DR": "{:,.2f}", "CR": "{:,.2f}", "Balance": "{:,.2f}"}, na_rep="")
        )

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            sheet_name = key if len(key) <= 31 else "Xns"
            styler.to_excel(writer, sheet_name=sheet_name, index=False)

            wb_out = writer.book
            ws     = writer.sheets[sheet_name]

            # ── header row ──────────────────────────────────────────────────
            for cell in ws[1]:
                cell.fill      = _HEADER_FILL
                cell.font      = _HEADER_FONT
                cell.border    = _THIN_BORDER
                cell.alignment = _CENTER_ALIGN

            # ── body rows ───────────────────────────────────────────────────
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border    = _THIN_BORDER
                    cell.font      = _BODY_FONT
                    cell.alignment = _VCENTER_ALIGN

            # ── column index map ─────────────────────────────────────────────
            col_index = {c.value: c.column for c in ws[1]}

            # Category column — bold Calibri
            cat_idx = col_index.get("Category")
            if cat_idx is not None:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                        min_col=cat_idx, max_col=cat_idx):
                    cell = row[0]
                    cell.font      = _CATEGORY_FONT
                    cell.alignment = Alignment(
                        wrap_text=cell.alignment.wrap_text if cell.alignment else False,
                        vertical="center",
                        horizontal=cell.alignment.horizontal if cell.alignment else "general",
                    )

            # Number columns — format all rows from row 1; Balance also checks negatives
            for col_name, check_neg in (("Balance", True), ("CR", False), ("DR", False)):
                idx = col_index.get(col_name)
                if idx is not None:
                    _apply_number_format_col(ws, idx, min_row=1, check_negative=check_neg)

            # Description column — wrap text
            desc_idx = col_index.get("Description")
            if desc_idx is not None:
                ws.cell(row=1, column=desc_idx).alignment = _WRAP_CENTER
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=desc_idx).alignment = _WRAP_VCENTER

            # Date column — date format
            date_idx = col_index.get("Date")
            if date_idx is not None:
                for r in range(2, ws.max_row + 1):
                    c = ws.cell(row=r, column=date_idx)
                    if c.value:
                        c.number_format = _DATE_FORMAT

            # Column widths + header row height
            for cell in ws[1]:
                w = _DESIRED_WIDTHS.get(cell.value)
                if w is not None:
                    wb_out[sheet_name].column_dimensions[cell.column_letter].width = w
            ws.row_dimensions[1].height = 20

            # ── Optional row highlighting ────────────────────────────────────
            if needs_highlight:
                red_pos   = (highlight_red_positions   or {}).get(key, set())
                green_pos = (highlight_green_positions or {}).get(key, set())
                n_rows    = len(df_out)
                for excel_row in range(2, ws.max_row + 1):
                    pos = excel_row - 2
                    if pos < 0 or pos >= n_rows:
                        continue
                    row_fill = None
                    if pos in green_pos:
                        row_fill = light_green_fill
                    if pos in red_pos:
                        row_fill = light_red_fill   # red overrides green
                    if row_fill is not None:
                        for cell in ws[excel_row]:
                            cell.fill = row_fill

            # ── Analysis + Statements sheets ─────────────────────────────────
            analysis_ws = copy_sheet_with_style(analysis_storage[key], wb_out, new_title="ANALYSIS")
            # copy_sheet_with_style already sets column widths; no need to repeat here
            _reorder_sheet_first(wb_out, analysis_ws)
            append_sheet_with_style(statement_storage[key], analysis_ws, gap_rows=3, col_offset=1)

        saved_files_info.append({
            'file_path':    str(filename),
            'file_name':    f"{safe_acc}-{safe_name}.xlsx",
            'account_name': acc_name,
            'sheet_name':   key,
        })
        print(f"✅ Saved styled file: {filename}")

    return saved_files_info
