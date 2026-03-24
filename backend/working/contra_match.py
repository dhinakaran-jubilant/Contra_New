from api import helpers
from api.regex_pattern import extract_imps
from api.inb_sis import infer_transfer_type
from api.compare_logic import find_imps_match, find_self_match, find_etxn_match, find_acc_num_match
from api.categorize_full import categorize_type
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter
from copy import copy
import pandas as pd

# ── Constants ──────────────────────────────────────────────────────────────────
_DESIRED_COL_WIDTHS = {
    "Sl. No.": 8, "Date": 12, "MONTH": 10, "TYPE": 12,
    "Cheque_No": 12, "Category": 35, "Description": 50,
    "DR": 15, "CR": 15, "Balance": 18,
}
_THRESHOLD = 10_00_000


# ── Module-level helpers (no closure needed) ──────────────────────────────────

def _preprocess_df(df, bank_name, file_type):
    """Preprocess DataFrame for comparison."""
    if "norm_date" not in df.columns:
        df["norm_date"] = df["Date"].apply(helpers.normalize_date) if "Date" in df.columns else pd.NaT

    for col, src in (("CR_val", "CR"), ("DR_val", "DR")):
        if col not in df.columns:
            df[col] = pd.to_numeric(df[src], errors="coerce").fillna(0.0).abs() if src in df.columns else 0.0

    if "IMPS" not in df.columns and "Description" in df.columns:
        df["IMPS"] = df["Description"].astype(str).apply(lambda d: extract_imps(d, bank_name))

    if "NUMBERS" not in df.columns and "Description" in df.columns:
        df["NUMBERS"] = df["Description"].astype(str).apply(helpers.get_numbers)

    if "IMPS" in df.columns and "NUMBERS" in df.columns:
        imps_pos = df.columns.get_loc("IMPS")
        nums = df.pop("NUMBERS")
        df.insert(imps_pos + 1, "NUMBERS", nums)

    return df


def _build_lookup_by_date(df, value_col):
    """Build date → [index, …] lookup for rows where value_col > 0."""
    if value_col not in df.columns or "norm_date" not in df.columns:
        return {}
    sub = df[df[value_col] > 0]
    lookup: dict = {}
    for idx, key in zip(sub.index, sub["norm_date"]):
        if pd.notna(key):
            lookup.setdefault(key, []).append(idx)
    return lookup


def _working_entry(sheet_name, info):
    """Build a file-info dict for a working file."""
    wb = info.get('workbook')
    return {
        'df':                    info['df'].copy(),
        'sheet_name':            sheet_name,
        'file_type':             'working',
        'acc_name':              info.get('acc_name', ''),
        'workbook':              wb,
        'worksheet':             wb[info['sheet_name']] if wb else None,
        'original_working_info': info,
    }


def _process_category(sheet_name: str) -> str:
    """Strip XNS segment from a sheet-name key."""
    parts = sheet_name.split("-")
    if "XNS" in parts:
        parts.remove("XNS")
    return "-".join(parts)


def _apply_cell_fmt(cell, fmt: dict) -> None:
    """Apply a stored format dict to *cell*."""
    if fmt.get('font'):      cell.font      = fmt['font']
    if fmt.get('fill'):      cell.fill      = fmt['fill']
    if fmt.get('border'):    cell.border    = fmt['border']
    if fmt.get('alignment'): cell.alignment = fmt['alignment']
    cell.number_format = fmt.get('number_format', 'General')


def _get_number_fmt(col_name: str | None, value, fallback: str = 'General') -> str:
    """Return the appropriate Excel number format string for *value* in *col_name*."""
    col_lower = str(col_name).lower() if col_name else ''
    if 'sl. no.' in col_lower:
        return '0'
    if 'date' in col_lower:
        return 'DD-MMM-YYYY' if isinstance(value, pd.Timestamp) else fallback
    if any(k in col_lower for k in ('dr', 'cr', 'balance', 'amount')):
        return '#,##0.00' if isinstance(value, (int, float)) else fallback
    if isinstance(value, pd.Timestamp):
        return 'DD-MMM-YYYY'
    if isinstance(value, (int, float)):
        return '0' if fallback == 'General' else fallback
    return fallback


# ── Main comparison engine ─────────────────────────────────────────────────────

def compare_all_files(software_data_storage, software_acc_name_storage,
                      working_data_storage, cash_deposit_sum=0, working_only=False):
    """
    Compare all files with each other.
    - working_only=False: software vs software, working vs working, software vs working
    - working_only=True : working vs working only
    """
    all_files = []

    if working_only:
        for name, info in working_data_storage.items():
            all_files.append(_working_entry(name, info))
    else:
        for name, df in software_data_storage.items():
            all_files.append({
                'df':         df.copy(),
                'sheet_name': name,
                'file_type':  'software',
                'acc_name':   software_acc_name_storage.get(name, ''),
            })
        for name, info in working_data_storage.items():
            all_files.append(_working_entry(name, info))

    print(f"Total files to process: {len(all_files)}")
    if working_only:
        print(f"Working files: {sum(1 for f in all_files if f['file_type'] == 'working')}")
    else:
        print(f"Software files: {sum(1 for f in all_files if f['file_type'] == 'software')}")
        print(f"Working files:  {sum(1 for f in all_files if f['file_type'] == 'working')}")

    total_matches = 0
    inb_trf_count = 0
    sis_con_count = 0
    match_details = []

    file_used_indices = {f['sheet_name']: set() for f in all_files}

    # Dispatch map: logic → find_*_match function
    _LOGIC_FN = {
        'imps':    find_imps_match,
        'etxn':    find_etxn_match,
    }

    def choose_candidate_for_row(row1, df2, lookup_df2, amount_col1, amount_col2,
                                 name1, name2, df1_sheet, df2_sheet, logic):
        try:
            if logic in _LOGIC_FN:
                idx2 = _LOGIC_FN[logic](row1, df2, lookup_df2, amount_col1, amount_col2,
                                        df2_sheet, file_used_indices)
            elif logic == 'self':
                idx2 = find_self_match(
                    row1=row1, df2=df2, lookup_df2=lookup_df2,
                    amount_col1=amount_col1, amount_col2=amount_col2,
                    this_acc_name=name1, other_acc_name=name2,
                    df1_key=df1_sheet, df2_key=df2_sheet,
                    file_used_indices=file_used_indices, is_working=True,
                )
            elif logic == 'acc_num':
                idx2 = find_acc_num_match(
                    row1, df2, lookup_df2, amount_col1, amount_col2,
                    df1_key=df1_sheet, df2_key=df2_sheet,
                    file_used_indices=file_used_indices,
                )
            else:
                return None
            return idx2 if (idx2 is not None and idx2 not in file_used_indices[df2_sheet]) else None
        except Exception as e:
            print(f'choose_candidate_for_row error: {e}')
            return None

    def apply_match(file1_info, file2_info, idx1, idx2, match_type, comparison_type):
        nonlocal total_matches, inb_trf_count, sis_con_count, match_details

        df1_sheet = file1_info['sheet_name']
        df2_sheet = file2_info['sheet_name']
        df1_type  = file1_info['file_type']
        df2_type  = file2_info['file_type']
        df1       = file1_info['df']
        df2       = file2_info['df']

        if idx1 in file_used_indices[df1_sheet] or idx2 in file_used_indices[df2_sheet]:
            return

        row2 = df2.iloc[idx2]

        if df2_type == 'working':
            working_type = str(row2.get('TYPE', '')).strip().upper()
            if 'UNMATCH INB TRF' in working_type or 'UNMAT INB TRF' in working_type or 'INB TRF' in match_type:
                final_type = "INB TRF"; inb_trf_count += 1
            elif 'UNMATCH SIS CON' in working_type or 'UNMAT SIS CON' in working_type or 'SIS CON' in match_type:
                final_type = "SIS CON"; sis_con_count += 1
            else:
                final_type = match_type
        else:
            final_type = match_type

        cat_df1 = _process_category(df1_sheet)
        cat_df2 = _process_category(df2_sheet)

        if df1_type == 'software':
            df1.at[idx1, "TYPE"]     = final_type
            df1.at[idx1, "Category"] = f"{file2_info['acc_name']}-{cat_df2}"

        if df2_type == 'software':
            df2.at[idx2, "TYPE"]     = final_type
            df2.at[idx2, "Category"] = f"{file1_info['acc_name']}-{cat_df1}"

        if df2_type == 'working' and file2_info.get('worksheet'):
            df2.at[idx2, "TYPE"]     = final_type
            df2.at[idx2, "Category"] = f"{file1_info['acc_name']}-{cat_df1}"
            update_working_file_new_columns(
                file2_info['worksheet'], idx2 + 2,
                final_type, f"{file1_info['acc_name']}-{cat_df1}",
            )

        if df1_type == 'working' and file1_info.get('worksheet'):
            df1.at[idx1, "TYPE"]     = final_type
            df1.at[idx1, "Category"] = f"{file2_info['acc_name']}-{cat_df2}"
            update_working_file_new_columns(
                file1_info['worksheet'], idx1 + 2,
                final_type, f"{file2_info['acc_name']}-{cat_df2}",
            )

        match_details.append({
            'file1':           df1_sheet,
            'file2':           df2_sheet,
            'file1_type':      df1_type,
            'file2_type':      df2_type,
            'file1_row':       idx1,
            'file2_row':       idx2,
            'comparison_type': comparison_type,
            'match_type':      final_type,
            'amount':          row2.get('DR_val') or row2.get('CR_val') or row2.get('Amount') or 0,
            'date':            row2.get('Date'),
        })
        file_used_indices[df1_sheet].add(idx1)
        file_used_indices[df2_sheet].add(idx2)
        total_matches += 1

    # ── Main comparison loop ──────────────────────────────────────────────────
    for logic in ('imps', 'self', 'etxn', 'acc_num'):
        print(f"\n{'='*60}\nUsing {logic.upper()} logic\n{'='*60}")

        for i, file1_info in enumerate(all_files):
            sheet1 = file1_info['sheet_name']
            bank1  = helpers.extract_bank_name_from_sheet(sheet1)
            file1_info['df'] = _preprocess_df(file1_info['df'], bank1, file1_info['file_type'])

            for file2_info in all_files[i + 1:]:
                sheet2 = file2_info['sheet_name']
                bank2  = helpers.extract_bank_name_from_sheet(sheet2)
                file2_info['df'] = _preprocess_df(file2_info['df'], bank2, file2_info['file_type'])

                print(f"\nComparing: {sheet1} ({file1_info['file_type']}) vs "
                      f"{sheet2} ({file2_info['file_type']})")

                if file1_info['file_type'] == file2_info['file_type'] == 'software':
                    cmp_type = 'software_vs_software'
                elif file1_info['file_type'] == file2_info['file_type'] == 'working':
                    cmp_type = 'working_vs_working'
                else:
                    cmp_type = 'software_vs_working'

                df1 = file1_info['df']
                df2 = file2_info['df']
                lookup_dr2 = _build_lookup_by_date(df2, "DR_val")
                lookup_cr2 = _build_lookup_by_date(df2, "CR_val")

                def _process_comparison(df_a, lookup_b, amount_col_a, amount_col_b,
                                        info_a, info_b, sheet_a, sheet_b):
                    for idx_a, row_a in df_a.iterrows():
                        if idx_a in file_used_indices[sheet_a]:
                            continue
                        idx_b = choose_candidate_for_row(
                            row_a, df2, lookup_b, amount_col_a, amount_col_b,
                            info_a['acc_name'], info_b['acc_name'],
                            sheet_a, sheet_b, logic,
                        )
                        if idx_b is not None:
                            apply_match(info_a, info_b, idx_a, idx_b,
                                        infer_transfer_type(info_a['acc_name'], info_b['acc_name']),
                                        cmp_type)

                _process_comparison(df1[df1["CR_val"] > 0], lookup_dr2, "CR_val", "DR_val",
                                    file1_info, file2_info, sheet1, sheet2)
                _process_comparison(df1[df1["DR_val"] > 0], lookup_cr2, "DR_val", "CR_val",
                                    file1_info, file2_info, sheet1, sheet2)

    # ── Clean up temp columns ─────────────────────────────────────────────────
    for file_info in all_files:
        df = file_info['df']
        df.drop(columns=[c for c in ("norm_date", "CR_val", "DR_val", "NUMBERS", "IMPS")
                         if c in df.columns], inplace=True)

        if working_only:
            working_data_storage[file_info['sheet_name']]['df'] = df
        else:
            type_value = 'SALES' if cash_deposit_sum > _THRESHOLD else 'CASH'
            acc_type   = file_info['sheet_name'].split('-')[-1]
            df = categorize_type(df, type_value, acc_type)
            if file_info['file_type'] == 'software':
                software_data_storage[file_info['sheet_name']] = df
            else:
                working_data_storage[file_info['sheet_name']]['df'] = df

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "="*60)
    print("FINAL SUMMARY")
    print("="*60)
    print(f"TOTAL MATCHES FOUND: {total_matches}")
    print(f"INB TRF COUNT: {inb_trf_count}")
    print(f"SIS CON COUNT: {sis_con_count}")

    by_type = {}
    for m in match_details:
        by_type[m['comparison_type']] = by_type.get(m['comparison_type'], 0) + 1

    if not working_only:
        print("\nComparison Breakdown:")
        print(f"  Software vs Software: {by_type.get('software_vs_software', 0)} matches")
        print(f"  Working vs Working:   {by_type.get('working_vs_working', 0)} matches")
        print(f"  Software vs Working:  {by_type.get('software_vs_working', 0)} matches")
    else:
        print(f"\nWorking vs Working matches: {by_type.get('working_vs_working', 0)}")

    return (working_data_storage, match_details) if working_only \
        else (software_data_storage, working_data_storage, match_details)


# ── Working-file Excel update ─────────────────────────────────────────────────

def update_working_file_new_columns(worksheet, row_num, type_value, category_value):
    """
    Write TYPE and Category into an existing working-file worksheet row,
    without touching column widths or other styles.
    """
    type_col = cat_col = None
    for col_idx, cell in enumerate(worksheet[1], start=1):
        header = str(cell.value).strip().upper() if cell.value else ""
        if header == "TYPE":
            type_col = col_idx
        elif header == "CATEGORY":
            cat_col = col_idx
        if type_col and cat_col:
            break

    if type_col is None or cat_col is None:
        return

    worksheet.cell(row=row_num, column=type_col).value     = type_value
    worksheet.cell(row=row_num, column=cat_col).value = category_value


# ── Save working files ────────────────────────────────────────────────────────

def save_working_files_with_styles(working_data_storage, output_dir_name="Matched_Statements"):
    """Save updated working files preserving original styles."""
    saved_files = []
    base_dir = helpers.get_downloads() / output_dir_name
    base_dir.mkdir(parents=True, exist_ok=True)

    for working_sheet_name, working_info in working_data_storage.items():
        try:
            wb          = working_info['workbook']
            updated_df  = working_info['df']
            orig_name   = working_info.get('original_file_name', working_sheet_name)
            acc_name    = working_info.get('acc_name', 'Working_File')
            sheet_name  = working_info['sheet_name']

            print(f"💾 Processing {sheet_name}...")

            # Find XNS sheet
            target_ws = orig_ws_name = None
            for ws_name in wb.sheetnames:
                if 'xns' in ws_name.lower():
                    target_ws    = wb[ws_name]
                    orig_ws_name = ws_name
                    print(f"  Found XNS sheet: {ws_name}")
                    break
            if not target_ws:
                print("⚠️ No XNS sheet found, using active sheet")
                target_ws    = wb.active
                orig_ws_name = target_ws.title

            # Capture header and data-row formatting templates
            print("  Storing original formatting...")

            def _capture_row_fmt(ws, row_idx):
                return {
                    col: {
                        'font':          copy(ws.cell(row=row_idx, column=col).font),
                        'fill':          copy(ws.cell(row=row_idx, column=col).fill),
                        'border':        copy(ws.cell(row=row_idx, column=col).border),
                        'alignment':     copy(ws.cell(row=row_idx, column=col).alignment),
                        'number_format': ws.cell(row=row_idx, column=col).number_format,
                    }
                    for col in range(1, ws.max_column + 1)
                }

            header_fmt   = _capture_row_fmt(target_ws, 1) if target_ws.max_row >= 1 else {}
            data_row_fmt = _capture_row_fmt(target_ws, 2) if target_ws.max_row >= 2 else {
                col: {'font': Font(), 'fill': PatternFill(fill_type=None),
                      'border': Border(), 'alignment': Alignment(),
                      'number_format': 'General'}
                for col in range(1, updated_df.shape[1] + 1)
            }

            # Store original column widths
            col_widths = {
                col: target_ws.column_dimensions[get_column_letter(col)].width
                for col in range(1, target_ws.max_column + 1)
                if get_column_letter(col) in target_ws.column_dimensions
            }

            # Clear cell values
            print("  Clearing sheet content...")
            for row in target_ws.iter_rows(max_row=target_ws.max_row, max_col=target_ws.max_column):
                for cell in row:
                    cell.value = None

            # Write data
            df_columns  = list(updated_df.columns)
            num_columns = len(df_columns)
            print(f"  Writing {len(updated_df)} records with original styles...")

            # Header row
            for ci, col_name in enumerate(df_columns, 1):
                cell = target_ws.cell(row=1, column=ci, value=col_name)
                if ci in header_fmt:
                    _apply_cell_fmt(cell, header_fmt[ci])
                else:
                    cell.font      = Font(bold=True, size=11)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Data rows
            for ri, row_data in enumerate(updated_df.itertuples(index=False), 2):
                for ci, value in enumerate(row_data, 1):
                    if ci > num_columns:
                        continue
                    cell     = target_ws.cell(row=ri, column=ci, value=value)
                    col_name = df_columns[ci - 1]
                    if ci in data_row_fmt:
                        _apply_cell_fmt(cell, data_row_fmt[ci])
                        cell.number_format = _get_number_fmt(
                            col_name, value, data_row_fmt[ci].get('number_format', 'General')
                        )
                    else:
                        cell.number_format = _get_number_fmt(col_name, value)

            # Column widths
            print("  Applying column widths...")
            for ci, col_name in enumerate(df_columns, 1):
                col_letter = get_column_letter(ci)
                if col_name in _DESIRED_COL_WIDTHS:
                    target_ws.column_dimensions[col_letter].width = _DESIRED_COL_WIDTHS[col_name]
                elif ci in col_widths:
                    target_ws.column_dimensions[col_letter].width = col_widths[ci]
                else:
                    max_len = max(
                        (len(str(c.value)) for c in target_ws[col_letter] if c.value),
                        default=0
                    )
                    target_ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

            target_ws.row_dimensions[1].height = 20

            output_filename = f"{acc_name}-{sheet_name}-WORKING-MERGED.xlsx"
            output_path     = base_dir / output_filename
            print(f"  Saving to {output_filename}...")
            wb.save(output_path)

            saved_files.append({
                'file_path':    str(output_path),
                'file_name':    output_filename,
                'account_name': acc_name,
                'sheet_name':   orig_ws_name,
                'original_file': orig_name,
                'record_count': len(updated_df),
                'is_merged':    working_info.get('merged_with_software', False),
            })
            print(f"✅ Saved: {output_filename} with {len(updated_df)} records\n")

        except Exception as e:
            print(f"❌ Error saving {working_sheet_name}: {e}")
            import traceback
            traceback.print_exc()

    print("\n" + "="*50)
    print(f"📁 SAVED {len(saved_files)} FILES TO: {base_dir}")
    for fi in saved_files:
        print(f"  • {fi['file_name']} ({fi['record_count']} records)")
    print("="*50)
    return saved_files


# ── Legacy / convenience wrappers ─────────────────────────────────────────────

def compare_software_with_working(software_data_storage, software_acc_name_storage, working_data_storage):
    """Legacy wrapper — use compare_all_files directly for new code."""
    print("Using new comprehensive comparison logic...")
    sw, wk, match_details = compare_all_files(
        software_data_storage, software_acc_name_storage, working_data_storage, working_only=False
    )
    filtered = [m for m in match_details if m['comparison_type'] == 'software_vs_working']
    return sw, filtered


def compare_and_save_files(software_data_storage, software_acc_name_storage,
                           working_data_storage, cash_deposit_sum=0, working_only=False):
    """Compare all files and optionally save updated working files."""
    if working_only:
        print("🔧 Running WORKING-ONLY comparison mode...")
        wk, match_details = compare_all_files(None, None, working_data_storage, working_only=True)
        saved = save_working_files_with_styles(wk, output_dir_name="Matched_Statements")
        return None, match_details, saved
    else:
        print("🔧 Running SOFTWARE + WORKING comparison mode...")
        sw, wk, match_details = compare_all_files(
            software_data_storage, software_acc_name_storage, working_data_storage,
            cash_deposit_sum, working_only=False,
        )
        return sw, wk, match_details


def compare_working_files_only(working_data_storage):
    """Compare working files only (no software files)."""
    print("🔧 Running WORKING-ONLY comparison...")
    wk, match_details = compare_all_files(None, None, working_data_storage, working_only=True)
    return wk, match_details
