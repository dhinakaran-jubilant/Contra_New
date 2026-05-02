import re
import traceback
import os
import mimetypes
import pandas as pd
import base64
import pythoncom
import win32com.client as win32
from pathlib import Path
from api.helpers import (
    get_sheet_name, get_month_values, log_processing, get_category_counts,
    count_inb_matches, count_return_matches, extract_bank_name_from_sheet,
    find_limit, extract_account_details_from_analysis, create_account_map_from_details,
    extract_account_suffix_from_sheet_name, get_acc_type, get_downloads,
    format_category_counts, sanitize_filename
)
from api import config
from api.contra_match import compare_files
from api.update_sheet import update_google_sheets
from api.style_helper import save_matched_with_styles
from api.categorize_full import categorize_desc_text, categorize_return_type, categorize_type
from api.pivot import create_pivot
from working.contra_match import compare_and_save_files, save_working_files_with_styles, compare_working_files_only
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import FileResponse, JsonResponse
from rest_framework.response import Response
from rest_framework.views import APIView
from openpyxl import load_workbook
from rest_framework import status

# ============================================================
# Module-level constants
# ============================================================
REQUIRED_COLS = ["Sl. No.", "Date", "MONTH", "TYPE", "Cheque_No", "Category", "Description", "DR", "CR", "Balance"]
MATCH_TYPES   = ["INB TRF", "SIS CON"]
CASH_THRESHOLD = 10_00_000


# ============================================================
# Module-level helper functions
# ============================================================

def _get_bank_name(sheet_name_or_code: str, use_code: bool = False) -> str:
    """Return the full bank name for a given sheet name or short bank code."""
    if use_code:
        for key, value in config.SHORT_BANK_NAMES.items():
            if value == sheet_name_or_code:
                return key
        return "Unknown Bank"
    bank_value = extract_bank_name_from_sheet(sheet_name_or_code)
    if bank_value:
        for key, value in config.SHORT_BANK_NAMES.items():
            if value == bank_value:
                return key
    return "Unknown Bank"


def _count_empty_type(df: pd.DataFrame) -> int:
    """Return the number of rows where the TYPE column is blank or NaN."""
    return int(((df['TYPE'].isna()) | (df['TYPE'].astype(str).str.strip() == '')).sum())


def _make_download_url(file_path: str) -> str:
    """Base64-encode a file path and return a download URL string."""
    encoded = base64.b64encode(str(file_path).encode()).decode()
    return f"/api/download-file/?file_path={encoded}"


# Remove local helpers that were moved to helpers.py


def _get_od_limit(df: pd.DataFrame, sheet_name: str):
    """Return the OD limit for an OD account, or None for non-OD accounts."""
    acc_type = str(sheet_name).split('-')[-1].strip()
    if acc_type != 'OD':
        return None
    negatives = df[df['Balance'] < 0]
    highest_negative = negatives['Balance'].min() if not negatives.empty else None
    return find_limit(highest_negative)


def _make_summary_entry(user_info: str, acc_name: str, sheet_name: str,
                        df: pd.DataFrame, matched: int, mode: str,
                        extra: dict = None) -> dict:
    """Build one summary report row."""
    safe_name = acc_name.replace("/", "_")
    empty_count = _count_empty_type(df)
    entry = {
        "User Name": user_info.upper(),
        "Bank Name": _get_bank_name(sheet_name),
        "File Name": f"{safe_name}-{sheet_name}",
        "Total Entries (Before)": len(df),
        "Contra Matches (Before)": matched,
        "Return (Before)": count_return_matches(df),
        "Total S/W Categorized": int(len(df) - empty_count),
    }
    if extra:
        entry.update(extra)
    return entry


def _make_software_download_entry(file_info: dict, df: pd.DataFrame, excel=None) -> dict:
    """Build a download-list entry for a software file, creating the pivot sheet."""
    file_path = file_info['file_path']
    sheet_name = file_info['sheet_name']
    limit = _get_od_limit(df, sheet_name)
    create_pivot(file_path, sheet_name, limit, excel=excel)
    return {
        "file_name": file_info['file_name'],
        "account_name": file_info['account_name'],
        "sheet_name": sheet_name,
        "file_type": "software",
        "file_size": os.path.getsize(file_path),
        "download_url": _make_download_url(file_path),
    }


def _make_working_download_entry(file_info: dict) -> dict:
    """Build a download-list entry for a working file."""
    file_path = file_info['file_path']
    return {
        "file_name": file_info['file_name'],
        "account_name": file_info.get('account_name', 'Working File'),
        "bank_name": file_info.get('bank_name', 'Unknown'),
        "account_suffix": file_info.get('account_suffix', ''),
        "xns_sheets": file_info.get('xns_sheets_count', 1),
        "file_type": "working",
        "file_size": os.path.getsize(file_path),
        "download_url": _make_download_url(file_path),
    }


def _merge_and_dedup(df_a: pd.DataFrame, df_b: pd.DataFrame,
                     cols: list = None) -> pd.DataFrame:
    """Concatenate two DataFrames, drop duplicates (ignoring Sl. No.),
    sort by Date, and reset the serial number."""
    cols = cols or REQUIRED_COLS
    combined = pd.concat([df_a, df_b], ignore_index=True)
    combined = combined.drop_duplicates(
        subset=[c for c in combined.columns if c != "Sl. No."],
        keep='first'
    )
    combined = combined.sort_values(by="Date", ascending=True).reset_index(drop=True)
    combined["Sl. No."] = combined.index + 1
    return combined[cols]


def _categorize_df(df: pd.DataFrame, bank_code: str = None) -> pd.DataFrame:
    """Apply description categorization and return-type categorization."""
    df['Category'] = df.apply(
        lambda row: categorize_desc_text(row['Description'], row['Category'], row['CR'], row['DR']),
        axis=1
    )
    if bank_code:
        df = categorize_return_type(df, bank_code)
    else:
        df = categorize_return_type(df)
    return df


def _parse_software_xns(file) -> tuple:
    """
    Read a software Excel file and return:
        (sheet_name, acc_name, acc_type, processed_df, wb_src)
    where processed_df has REQUIRED_COLS but WITHOUT categorization applied yet.
    """
    analysis_df = pd.read_excel(file, sheet_name="Analysis", header=None)
    sheet_name, acc_name = get_sheet_name(analysis_df)

    xns_df = pd.read_excel(file, sheet_name="Xns")
    acc_type = get_acc_type(xns_df["Balance"])
    sheet_name = f"{sheet_name}-{acc_type}"

    xns_df["Amount"] = (
        xns_df["Amount"].astype(str).str.replace(",", "", regex=False).astype(float)
    )
    xns_df["DR"] = xns_df["Amount"].where(xns_df["Type"] == "Debit")
    xns_df["CR"] = xns_df["Amount"].where(xns_df["Type"] == "Credit")
    xns_df["Type"] = ""

    new_df = xns_df.drop(columns=["Amount"]).rename(columns={
        "Sl. No. ": "Sl. No.",
        "Cheque No.": "Cheque_No",
        "Type": "TYPE",
    })
    new_df['MONTH'] = get_month_values(new_df, "Date")
    new_df["Date"] = pd.to_datetime(new_df["Date"], errors="coerce")
    new_df = new_df[REQUIRED_COLS]

    wb_src = load_workbook(file, data_only=False)
    return sheet_name, acc_name, acc_type, new_df, wb_src


# ============================================================
# MatchStatement view
# ============================================================

class MatchStatement(APIView):
    parser_classes = [MultiPartParser, FormParser]

    # ----------------------------------------------------------
    # Entry point
    # ----------------------------------------------------------
    def post(self, request, *args, **kwargs):
        # Get user_name from request data (passed from frontend) or fallback to authenticated user
        user_info = request.data.get('user_name')
        if not user_info:
            user_info = request.user.username if request.user and request.user.is_authenticated else "Anonymous"
        
        excel_files = request.FILES.getlist("files")

        if not excel_files:
            return Response(
                {"error": "No files uploaded. Please upload .xlsx files."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate extensions
        valid_files = []
        for f in excel_files:
            name = str(f.name)
            if Path(name).suffix.lower() != ".xlsx":
                return Response(
                    {"error": "Only .xlsx files are allowed.", "invalid_file": name},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            valid_files.append(f)

        print(f"📁 Total files uploaded: {len(valid_files)}")

        # Auto-detect file types
        software_files, working_files = [], []
        for file in valid_files:
            try:
                xl = pd.ExcelFile(file)
                if 'Xns' in xl.sheet_names:
                    software_files.append(file)
                    print(f"✅ Software file (has 'Xns' sheet): {file.name}")
                else:
                    working_files.append(file)
                    print(f"📝 Working file (no 'Xns' sheet): {file.name}")
            except Exception as e:
                print(f"⚠️ Could not determine type for {file.name}, treating as working: {e}")
                working_files.append(file)

        n_sw, n_wk = len(software_files), len(working_files)
        print(f"📊 Classification — Software: {n_sw}, Working: {n_wk}")

        if n_sw >= 2 and n_wk == 0:
            print("🔧 Mode: Software-only comparison")
            return self._handle_software_only_comparison(software_files, user_info, request)
        elif n_sw >= 1 and n_wk >= 1:
            print("🔧 Mode: Software-Working comparison")
            return self._handle_software_working_comparison(software_files, working_files, user_info, request)
        elif n_sw == 1 and n_wk == 0:
            print("🔧 Mode: Single software file preprocessing")
            return self._handle_single_software_file(software_files[0], user_info, request)
        elif n_sw == 0 and n_wk >= 2:
            print("🔧 Mode: Working-only comparison")
            return self._handle_working_only_comparison(working_files, user_info, request)
        elif n_sw == 0 and n_wk == 1:
            return Response(
                {"error": f"At least 2 working files are required for working-only comparison. Only {n_wk} found."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        else:
            return Response(
                {
                    "error": (
                        "Invalid file combination. Need either:\n"
                        "- 2+ software files (software-only comparison)\n"
                        "- 1+ software AND 1+ working files (software-working comparison)\n"
                        "- 1 software file (preprocessing only)\n"
                        "- 2+ working files (working-only comparison)"
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

    # ----------------------------------------------------------
    # Load working file(s) — shared private helper
    # ----------------------------------------------------------
    def _load_working_file_detailed(self, file) -> dict:
        """
        Parse a working Excel file and return a dict keyed by XNS sheet name.
        Handles multiple XNS sheets and full account-details resolution.
        """
        wb = load_workbook(file, data_only=False)

        # Extract account details from ANALYSIS sheet
        account_details_map = {}
        try:
            if "ANALYSIS" in wb.sheetnames:
                analysis_df = pd.read_excel(file, sheet_name="ANALYSIS", header=None)
                account_details_list = extract_account_details_from_analysis(analysis_df)
                account_details_map = create_account_map_from_details(account_details_list)
                print(f"📊 Found {len(account_details_map)} accounts in ANALYSIS sheet")
            else:
                print(f"ℹ️ No ANALYSIS sheet in {file.name}")
        except Exception as e:
            print(f"⚠️ Could not parse ANALYSIS sheet: {e}")
            traceback.print_exc()

        xns_sheet_names = [s for s in wb.sheetnames if 'xns' in s.lower()]
        if not xns_sheet_names:
            print(f"⚠️ No XNS sheet in file: {file.name}")
            return {}

        print(f"📊 Found {len(xns_sheet_names)} XNS sheets: {xns_sheet_names}")

        result = {}
        for xns_name in xns_sheet_names:
            try:
                account_suffix = extract_account_suffix_from_sheet_name(xns_name)
                if not account_suffix:
                    matches = re.findall(r'\d{4,5}', xns_name)
                    if matches:
                        account_suffix = matches[-1]
                        print(f"ℹ️ Suffix via regex: {account_suffix}")

                if not account_suffix:
                    acc_name, bank_name, bank_code = xns_name, "Unknown Bank", None
                    account_info = None
                else:
                    if len(account_suffix) == 3:
                        account_suffix = f"X{account_suffix}"
                    account_info = account_details_map.get(account_suffix)
                    if account_info:
                        acc_name  = account_info['account_holder']
                        bank_name = account_info['bank_name']
                        bank_code = account_info['bank_code']
                    else:
                        bank_code = None
                        sheet_upper = xns_name.upper()
                        for bank_key, bank_short in config.SHORT_BANK_NAMES.items():
                            if bank_short in sheet_upper or bank_key.upper() in sheet_upper:
                                bank_code = bank_short
                                bank_name = bank_key
                                break
                        if not bank_code:
                            for part in sheet_upper.split('-'):
                                if part in config.SHORT_BANK_NAMES.values():
                                    bank_code = part
                                    bank_name = _get_bank_name(part, use_code=True)
                                    break
                        if bank_code:
                            acc_name = f"{bank_name}-{account_suffix}"
                        else:
                            acc_name, bank_name = f"Account-{account_suffix}", "Unknown Bank"

                df = pd.read_excel(file, sheet_name=xns_name)
                result[xns_name] = {
                    'df': df,
                    'sheet_name': xns_name,
                    'workbook': wb,
                    'acc_name': acc_name,
                    'bank_code': bank_code,
                    'account_suffix': account_suffix,
                    'bank_name': bank_name,
                    'workbook_instance': wb,
                    'file_object': file,
                    'account_info': account_info or {},
                    'original_file_name': file.name,
                }
                print(f"✅ Loaded {acc_name} | Sheet: {xns_name} | Bank: {bank_name} | Rows: {len(df)}")
            except Exception as e:
                print(f"❌ Error processing XNS sheet {xns_name}: {e}")
                traceback.print_exc()
        return result

    # ----------------------------------------------------------
    # Handler: single software file
    # ----------------------------------------------------------
    def _handle_single_software_file(self, software_file, user_info, request):
        """Preprocess a single software file — no cross-account comparison."""
        try:
            print(f"📂 Single software file: {software_file.name}")
            sheet_name, acc_name, acc_type, new_df, wb_src = _parse_software_xns(software_file)

            new_df = _categorize_df(new_df, bank_code=sheet_name.split('-')[1].strip())
            new_df = categorize_return_type(new_df, sheet_name.split('-')[1].strip())

            cash_deposit_sum = new_df.loc[new_df['Category'] == 'CASH DEPOSIT', 'CR'].sum()
            type_value = 'SALES' if cash_deposit_sum > CASH_THRESHOLD else 'CASH'
            new_df = categorize_type(new_df, type_value, acc_type)
            print(f"✅ Categorized: {sheet_name}")

            acc_name_storage  = {sheet_name: acc_name}
            analysis_storage  = {sheet_name: wb_src["Analysis"]}
            statement_storage = {sheet_name: wb_src["Statements Considered"]}

            saved_files_info = save_matched_with_styles(
                {sheet_name: new_df}, acc_name_storage, analysis_storage, statement_storage
            )
            if not saved_files_info:
                return Response(
                    {"error": "Failed to save the preprocessed file."},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

            saved = saved_files_info[0]
            file_path = saved['file_path']

            download_files = []
            if os.path.exists(file_path):
                # We can just use the default create_pivot behavior here since it's only one file,
                # but for consistency with the optimized flow:
                download_files.append(_make_software_download_entry(saved, new_df))

            summary = [_make_summary_entry(
                user_info, acc_name, sheet_name, new_df,
                matched=0, mode="Single File Preprocessing"
            )]

            total_entries = len(new_df)
            
            # Log processing
            log_processing(
                user_name=user_info,
                file_name=summary[0]['File Name'],
                bank_name=_get_bank_name(sheet_name),
                total_entries=total_entries,
                software_count=format_category_counts(new_df),
                final_count=""
            )

            # Update Google Sheets
            try:
                print("📊 Updating Google Sheets (Single File)...")
                update_google_sheets(summary, is_live=True)
            except Exception as gs_err:
                print(f"⚠️ Google Sheets update failed: {gs_err}")

            return Response({
                "success": True,
                "message": f"Single software file preprocessed successfully. Saved as {saved['file_name']}.",
                "comparison_mode": "single-file-preprocess",
                "files_processed": 1,
                "total_entries": total_entries,
                "summary": summary,
                "download_files": download_files,
                "total_files_to_download": len(download_files),
                "output_directory": str(Path(file_path).parent),
            }, status=status.HTTP_200_OK)

        except Exception as e:
            print(f"❌ Single software file error: {e}")
            traceback.print_exc()
            return Response(
                {"error": f"Error processing file: {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    # ----------------------------------------------------------
    # Handler: working-only comparison
    # ----------------------------------------------------------
    def _handle_working_only_comparison(self, working_files, user_info, request):
        """Compare 2+ working files against each other."""
        try:
            print(f"🔧 Working-only: {len(working_files)} files")
            working_data_storage = {}

            for file in working_files:
                try:
                    sheets = self._load_working_file_detailed(file)
                    working_data_storage.update(sheets)
                except Exception as e:
                    print(f"❌ Error loading {file.name}: {e}")
                    traceback.print_exc()

            if len(working_data_storage) < 2:
                return Response(
                    {"error": f"At least 2 XNS sheets required. Only {len(working_data_storage)} loaded."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            print(f"\n{'='*60}\nRUNNING WORKING-WORKING COMPARISON\n{'='*60}")
            try:
                _, match_details = compare_working_files_only(working_data_storage)
            except Exception as e:
                traceback.print_exc()
                return Response(
                    {"error": f"Error during comparison: {e}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

            saved_working_files_info = save_working_files_with_styles(
                working_data_storage, output_dir_name="Matched_Statements"
            )

            summary_report = []
            download_files = []
            for xns_name, info in working_data_storage.items():
                df           = info['df']
                acc_name     = info['acc_name']
                account_suffix = info.get('account_suffix', '')
                bank_code    = info.get('bank_code')
                full_bank    = _get_bank_name(bank_code, use_code=True) if bank_code else info.get('bank_name', 'Unknown Bank')
                empty_count  = _count_empty_type(df)
                safe_acc_name = sanitize_filename(acc_name)
                file_name = f"{safe_acc_name}-{account_suffix}-WORKING"
                summary_report.append({
                    "User Name": user_info.upper(),
                    "Bank Name": full_bank,
                    "File Name": file_name,
                    "Total Entries (Before)": len(df),
                    "Contra Matches (Before)": count_inb_matches(df),
                    "Return (Before)": count_return_matches(df),
                    "Total S/W Categorized": int(len(df) - empty_count),
                })
                
                # Log processing
                log_processing(
                    user_name=user_info,
                    file_name=file_name,
                    bank_name=full_bank,
                    total_entries=len(df),
                    software_count=format_category_counts(df),
                    final_count=""
                )

            # Update Google Sheets
            try:
                print("📊 Updating Google Sheets (Working Only)...")
                update_google_sheets(summary_report, is_live=True)
            except Exception as gs_err:
                print(f"⚠️ Google Sheets update failed: {gs_err}")

            total_entries = sum(len(info['df']) for info in working_data_storage.values())

            for fi in saved_working_files_info:
                if os.path.exists(fi['file_path']):
                    download_files.append(_make_working_download_entry(fi))

            return Response({
                "success": True,
                "message": (
                    f"Working-only comparison completed. "
                    f"Found {len(match_details)} matches across {len(working_data_storage)} XNS sheets. "
                    f"Updated {len(saved_working_files_info)} working files."
                ),
                "comparison_mode": "working-only",
                "user_info": user_info.upper(),
                "total_entries": total_entries,
                "summary": {
                    "total_matches": len(match_details),
                    "xns_sheets_processed": len(working_data_storage),
                    "unique_accounts": len(set(i['acc_name'] for i in working_data_storage.values())),
                    "files_saved": len(saved_working_files_info),
                },
                "match_details": match_details[:50],
                "download_files": download_files,
                "total_files_to_download": len(download_files),
                "output_directory": str(Path(saved_working_files_info[0]['file_path']).parent) if saved_working_files_info else "N/A",
            }, status=status.HTTP_200_OK)

        except Exception as e:
            print(f"❌ Working-only error: {e}")
            traceback.print_exc()
            return Response(
                {"error": f"Error in working-only comparison: {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    # ----------------------------------------------------------
    # Handler: software-only comparison
    # ----------------------------------------------------------
    def _handle_software_only_comparison(self, software_files, user_info, request):
        """Compare 2+ software files against each other."""
        bank_data_storage = {}
        analysis_storage  = {}
        acc_name_storage  = {}
        statement_storage = {}
        cash_deposit_sum  = 0
        analysis_count    = 1

        software_counts_before = {}
        for file in software_files:
            try:
                print(f"📂 Software file: {file.name}")
                sheet_name, acc_name, acc_type, new_df, wb_src = _parse_software_xns(file)

                # Capture counts before any matching/merging
                software_counts_before[sheet_name] = format_category_counts(new_df)

                if sheet_name in bank_data_storage:
                    print(f"⚠️ Merging duplicate sheet: {sheet_name}")
                    existing_df = bank_data_storage[sheet_name]
                    for col in REQUIRED_COLS:
                        if col not in existing_df.columns:
                            existing_df[col] = None
                    cash_deposit_sum += existing_df.loc[existing_df['Category'] == 'CASH DEPOSIT', 'CR'].sum()
                    combined_df = _merge_and_dedup(existing_df[REQUIRED_COLS], new_df)
                    combined_df = _categorize_df(combined_df)
                    cash_deposit_sum += combined_df.loc[combined_df['Category'] == 'CASH DEPOSIT', 'CR'].sum()
                    bank_data_storage[sheet_name] = combined_df
                else:
                    new_df = _categorize_df(new_df)
                    cash_deposit_sum += new_df.loc[new_df['Category'] == 'CASH DEPOSIT', 'CR'].sum()
                    bank_data_storage[sheet_name] = new_df

                acc_name_storage[sheet_name] = acc_name

                key_analysis   = sheet_name if sheet_name not in analysis_storage  else f"{sheet_name}_{analysis_count}"
                key_statement  = sheet_name if sheet_name not in statement_storage else f"{sheet_name}_{analysis_count}"
                analysis_storage[key_analysis]   = wb_src["Analysis"]
                statement_storage[key_statement] = wb_src["Statements Considered"]
                if key_analysis != sheet_name:
                    analysis_count += 1

                print(f"✅ Loaded: {sheet_name}")
            except Exception as e:
                print(f"❌ Software file error {file.name}: {e}")
                traceback.print_exc()

        if len(bank_data_storage) < 2:
            return Response(
                {"error": f"At least 2 software files required. Only {len(bank_data_storage)} loaded."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        print(f"📁 Loaded {len(bank_data_storage)} files: {list(bank_data_storage.keys())}")

        matched_df = compare_files(
            bank_data_storage=bank_data_storage,
            acc_name_storage=acc_name_storage,
            cash_deposit_sum=cash_deposit_sum,
        )

        all_files_to_process = {
            key: matched_df.get(key, bank_data_storage[key])
            for key in bank_data_storage
        }

        saved_files_info = save_matched_with_styles(
            all_files_to_process, acc_name_storage, analysis_storage, statement_storage
        )

        summary_report = [
            _make_summary_entry(
                user_info,
                acc_name_storage.get(sn, ""),
                sn, df,
                matched=count_inb_matches(df),
                mode="Software Only",
            )
            for sn, df in all_files_to_process.items()
        ]

        for (sn, df), report_entry in zip(all_files_to_process.items(), summary_report):
            log_processing(
                user_name=user_info,
                file_name=report_entry['File Name'],
                bank_name=_get_bank_name(sn),
                total_entries=len(df),
                software_count=format_category_counts(df),
                final_count=""
            )

        # Update Google Sheets
        try:
            print("📊 Updating Google Sheets (Software Only)...")
            update_google_sheets(summary_report, is_live=True)
        except Exception as gs_err:
            print(f"⚠️ Google Sheets update failed: {gs_err}")

        total_entries = sum(len(df) for df in all_files_to_process.values())

        # --- Build download list (Batch pivot generation) ---
        excel = None
        try:
            pythoncom.CoInitialize()
            excel = win32.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            try: excel.Calculation = -4135 # xlCalculationManual
            except: pass

            download_files = [
                _make_software_download_entry(fi, all_files_to_process[fi['sheet_name']], excel=excel)
                for fi in saved_files_info
                if os.path.exists(fi['file_path'])
            ]
        finally:
            if excel:
                try: excel.Calculation = -4105
                except: pass
                try: excel.Quit()
                except: pass
            pythoncom.CoUninitialize()

        return Response({
            "success": True,
            "message": f"Software-only matching completed. Processed {len(all_files_to_process)} software files.",
            "comparison_mode": "software-only",
            "files_processed": len(all_files_to_process),
            "total_entries": total_entries,
            "files_in_summary": len(summary_report),
            "summary": summary_report,
            "download_files": download_files,
            "total_files_to_download": len(download_files),
            "output_directory": str(get_downloads() / "Matched_Statements"),
        }, status=status.HTTP_200_OK)

    # ----------------------------------------------------------
    # Handler: software + working comparison
    # ----------------------------------------------------------
    def _handle_software_working_comparison(self, software_files, working_files, user_info, request):
        """Compare software files against working files."""
        # --- Load working files (simple: first XNS sheet only) ---
        working_data_storage = {}
        cash_deposit_sum = 0

        for file in working_files:
            try:
                wb = load_workbook(file, data_only=False)
                xns_sheet_name = next(
                    (s for s in wb.sheetnames if 'xns' in s.lower()), None
                )
                if not xns_sheet_name:
                    print(f"⚠️ No XNS sheet in {file.name}")
                    continue

                try:
                    analysis_df = pd.read_excel(file, sheet_name="ANALYSIS", header=None)
                    _, acc_name = get_sheet_name(analysis_df)
                except Exception:
                    acc_name = "Working File"

                df = pd.read_excel(file, sheet_name=xns_sheet_name)
                cash_deposit_sum += df.loc[df['Category'].str.upper() == 'CASH DEPOSIT', 'CR'].sum()
                working_data_storage[xns_sheet_name] = {
                    'df': df, 'sheet_name': xns_sheet_name,
                    'workbook': wb, 'acc_name': acc_name,
                    'original_file_name': file.name,
                }
                print(f"✅ Working file loaded: {file.name} (Sheet: {xns_sheet_name})")
            except Exception as e:
                print(f"❌ Error loading working file {file.name}: {e}")

        if not working_data_storage:
            return Response(
                {"error": "At least 1 valid working file is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # --- Load software files ---
        software_data_storage   = {}
        software_acc_name_storage = {}
        analysis_storage        = {}
        statement_storage       = {}

        software_counts_before = {}
        for file in software_files:
            try:
                sheet_name, acc_name, acc_type, new_df, wb_src = _parse_software_xns(file)
                
                # Capture counts before any matching/merging
                software_counts_before[sheet_name] = format_category_counts(new_df)

                # Check if this software sheet matches any working sheet
                matching_working_sheet = None
                for wk_name in working_data_storage:
                    wk_norm = wk_name.upper()
                    for sfx in ('-CA', 'CA', '-OD', 'OD', '-SB', 'SB'):
                        wk_norm = wk_norm.replace(sfx, '')
                    wk_norm = wk_norm.strip('-').strip()
                    if wk_norm == sheet_name.replace(acc_type, '').strip('-'):
                        matching_working_sheet = wk_name
                        print(f"🔍 Software '{sheet_name}' matches working '{wk_name}'")
                        break

                if matching_working_sheet:
                    # Merge software data into the matching working entry
                    wk_df = working_data_storage[matching_working_sheet]['df'].copy()
                    cash_deposit_sum -= wk_df.loc[wk_df['Category'].str.upper() == 'CASH DEPOSIT', 'CR'].sum()
                    wk_df['Date'] = pd.to_datetime(wk_df['Date'], errors='coerce')
                    merged = _merge_and_dedup(wk_df, new_df.copy())
                    merged = _categorize_df(merged)
                    cash_deposit_sum += merged.loc[merged['Category'].str.upper() == 'CASH DEPOSIT', 'CR'].sum()
                    working_data_storage[matching_working_sheet]['df'] = merged
                    print(f"✅ Merged {len(new_df)} rows into '{matching_working_sheet}' → {len(merged)} total")

                elif sheet_name in software_data_storage:
                    existing_df = software_data_storage[sheet_name]
                    for col in REQUIRED_COLS:
                        if col not in existing_df.columns:
                            existing_df[col] = None
                    cash_deposit_sum -= existing_df.loc[existing_df['Category'].str.upper() == 'CASH DEPOSIT', 'CR'].sum()
                    combined = _merge_and_dedup(existing_df[REQUIRED_COLS], new_df)
                    combined = _categorize_df(combined)
                    cash_deposit_sum += combined.loc[combined['Category'].str.upper() == 'CASH DEPOSIT', 'CR'].sum()
                    software_data_storage[sheet_name] = combined

                else:
                    new_df = _categorize_df(new_df)
                    cash_deposit_sum += new_df.loc[new_df['Category'].str.upper() == 'CASH DEPOSIT', 'CR'].sum()
                    software_data_storage[sheet_name] = new_df

                software_acc_name_storage[sheet_name] = acc_name
                analysis_storage[sheet_name]  = wb_src["Analysis"]
                statement_storage[sheet_name] = wb_src["Statements Considered"]
                print(f"✅ Software loaded: {sheet_name}")

            except Exception as e:
                print(f"❌ Software file error {file.name}: {e}")
                traceback.print_exc()

        if not software_data_storage:
            return Response(
                {"error": "At least 1 valid software file is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # --- Run comparison ---
        print("\n=== RUNNING SOFTWARE-WORKING COMPARISON ===")
        try:
            updated_software_data, updated_working_data, match_details = compare_and_save_files(
                software_data_storage, software_acc_name_storage,
                working_data_storage, float(cash_deposit_sum)
            )
        except Exception as e:
            traceback.print_exc()
            return Response(
                {"error": f"Error during comparison: {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        # --- Save files ---
        print("\n=== SAVING UPDATED FILES ===")
        saved_working_files_info = save_working_files_with_styles(
            updated_working_data, output_dir_name="Matched_Statements"
        )
        saved_software_files_info = save_matched_with_styles(
            updated_software_data, software_acc_name_storage, analysis_storage, statement_storage
        )

        # --- Build summary ---
        summary_report = [
            _make_summary_entry(
                user_info, software_acc_name_storage.get(sn, ""), sn, df,
                matched=count_inb_matches(df), mode="Software-Working"
            )
            for sn, df in updated_software_data.items()
        ]
        for wk_name, wk_info in working_data_storage.items():
            df        = wk_info['df']
            acc_name  = wk_info.get('acc_name', 'Working File')
            safe_name = acc_name.replace("/", "_")
            empty_c   = _count_empty_type(df)
            summary_report.append({
                "User Name": user_info.upper(),
                "Bank Name": _get_bank_name(wk_name),
                "File Name": f"{safe_name}-{wk_name}-WORKING",
                "Total Entries (Before)": len(df),
                "Contra Matches (Before)": count_inb_matches(df),
                "Return (Before)": count_return_matches(df),
                "Total S/W Categorized": int(len(df) - empty_c),
            })

        # Update Google Sheets
        try:
            print("📊 Updating Google Sheets (Software-Working)...")
            update_google_sheets(summary_report, is_live=True)
        except Exception as gs_err:
            print(f"⚠️ Google Sheets update failed: {gs_err}")

        for (sn, df), report_entry in zip(updated_software_data.items(), summary_report):
            log_processing(
                user_name=user_info,
                file_name=report_entry['File Name'],
                bank_name=_get_bank_name(sn),
                total_entries=len(df),
                software_count=format_category_counts(df),
                final_count=""
            )
        
        # Log processing for working files (offset by the software entries in summary_report)
        sw_count = len(updated_software_data)
        for i, (wk_name, wk_info) in enumerate(working_data_storage.items()):
            df = wk_info['df']
            report_entry = summary_report[sw_count + i]
            bank_code = wk_info.get('bank_code')
            full_bank = _get_bank_name(bank_code, use_code=True) if bank_code else wk_info.get('bank_name', 'Unknown Bank')
            log_processing(
                user_name=user_info,
                file_name=report_entry['File Name'],
                bank_name=full_bank,
                total_entries=len(df),
                software_count=format_category_counts(df),
                final_count=""
            )

        # --- Build download list (Batch pivot generation) ---
        excel = None
        try:
            pythoncom.CoInitialize()
            excel = win32.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            try: excel.Calculation = -4135 # xlCalculationManual
            except: pass

            download_files = [
                _make_working_download_entry(fi)
                for fi in saved_working_files_info
                if os.path.exists(fi['file_path'])
            ] + [
                _make_software_download_entry(fi, updated_software_data[fi['sheet_name']], excel=excel)
                for fi in saved_software_files_info
                if os.path.exists(fi['file_path'])
            ]
        finally:
            if excel:
                try: excel.Calculation = -4105
                except: pass
                try: excel.Quit()
                except: pass
            pythoncom.CoUninitialize()

        # Calculate total entries
        total_软件 = sum(len(df) for df in updated_software_data.values())
        total_working = sum(len(info['df']) for info in updated_working_data.values())
        total_entries = int(total_软件 + total_working)

        return Response({
            "success": True,
            "message": (
                f"Software-Working comparison completed. "
                f"Processed {len(updated_software_data)} software and {len(updated_working_data)} working files."
            ),
            "comparison_mode": "software-working",
            "files_processed": len(software_files) + len(working_files),
            "total_entries": total_entries,
            "user_info": user_info.upper(),
            "summary": {
                "total_matches": len(match_details),
                "inb_trf_matches": sum(1 for m in match_details if m['match_type'] == 'INB TRF'),
                "sis_con_matches": sum(1 for m in match_details if m['match_type'] == 'SIS CON'),
                "software_files_processed": len(software_data_storage),
                "working_files_processed": len(working_data_storage),
                "files_saved": len(saved_working_files_info) + len(saved_software_files_info),
            },
            "match_details": match_details[:50],
            "download_files": download_files,
            "total_files_to_download": len(download_files),
            "output_directory": str(get_downloads() / "Matched_Statements"),
        }, status=status.HTTP_200_OK)


# ============================================================
# DownloadFileView
# ============================================================

class DownloadFileView(APIView):
    """Serve individual processed files as downloads."""

    def get(self, request):
        """Download file endpoint."""
        # Cache allowed dirs for performance
        if not hasattr(self.__class__, '_allowed_dirs'):
            downloads = get_downloads()
            self.__class__._allowed_dirs = [
                (downloads / "Matched_Statements").resolve(),
                (downloads / config.PROCESSED_DIR).resolve(),
            ]

        try:
            file_path_encoded = request.GET.get('file_path')
            if not file_path_encoded:
                return Response({"error": "No file path provided"}, status=status.HTTP_400_BAD_REQUEST)

            file_path_str = base64.b64decode(file_path_encoded.encode()).decode()
            file_path = Path(file_path_str)
            
            # Efficient path check using is_relative_to
            file_allowed = any(
                file_path.is_relative_to(d)
                for d in self.__class__._allowed_dirs
            )
            if not file_allowed:
                return JsonResponse(
                    {"error": "Access to this file is not allowed"},
                    status=status.HTTP_403_FORBIDDEN,
                )
            if not os.path.exists(file_path):
                return JsonResponse(
                    {"error": "File not found"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            file_name    = os.path.basename(file_path)

            # If this is a consolidated file, trigger Google Sheets update asynchronously
            if "-CONSOLIDATED" in file_name.upper():
                try:
                    import threading
                    from api.update_sheet import update_google_sheets_final
                    print(f"🚀 Triggering ASYNC Google Sheets update for: {file_name}")
                    threading.Thread(target=update_google_sheets_final, args=(str(file_path),), daemon=True).start()
                except Exception as gs_err:
                    print(f"⚠️ Failed to trigger async GS update: {gs_err}")
            content_type, _ = mimetypes.guess_type(file_path)
            content_type = content_type or 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            response = FileResponse(open(file_path, 'rb'), content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['Pragma']  = 'no-cache'
            response['Expires'] = '0'
            return response

        except Exception as e:
            print(f"❌ Download error: {e}")
            return JsonResponse(
                {"error": f"Error downloading file: {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
